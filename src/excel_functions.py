"""
This file contains functions for working with Excel files.
The functions in this file are called by xl_to_rahshal.py
"""


import openpyxl
from docx import Document


def find_area_in_xl(nz_xl, xl_row: int) -> [str, int]:
    """
    The function `find_area_in_xl` searches for the first non-empty cell value in column A of an Excel
    file starting from a specified row, and returns the cell value and the row number where it was
    found.

    Args:
    nz_xl: The variable `nz_xl` represents an Excel file or workbook. It is not clear from the code
    snippet how this variable is defined or obtained, but it is likely an object that provides access to
    the contents of an Excel file.
    xl_row (int): The `xl_row` parameter is an integer that represents the row number in the Excel
    file from which you want to start searching for the area.

    Returns:
    The function `find_area_in_xl` returns a tuple containing a string and an integer. The string
    represents the cell value found in column A, and the integer represents the row number where the
    cell value was found.
    """
    # Starting from one row after xl_row (see docs for xl_row) until the the last row in te excel file.
    # I added +1 to max_row because otherwise it will go upto the one to last but not the last row
    for row in range(xl_row + 1, nz_xl.max_row + 1):
        # Get the cell value of the cell in column A
        cell_value: str = nz_xl[f"A{row}"].value

        if cell_value is not None and "קורדינטות" not in cell_value:
            # Returns the cell value and the row number where the cell value was found.
            return cell_value, row

    # If the code reaches this point, then the excel file is exhausted and no more areas are found.
    return " ", row


def xl_table_dimensions(nz_xl, xl_row: int) -> int:
    """
    The function `xl_table_dimensions` takes in an Excel worksheet object `nz_xl` and an integer
    `xl_row`, and returns the number of rows in a table starting from `xl_row` in the worksheet.

    Args:
    nz_xl: The parameter `nz_xl` is expected to be an object representing an Excel workbook or
    worksheet. It is used to access the cells in the worksheet and retrieve their values.
    xl_row (int): The `xl_row` parameter represents the starting row from which you want to find the
    dimensions of the table.

    Returns:
    The function `xl_table_dimensions` returns the number of rows in a table, calculated as the
    difference between the last row and the first row of the table.
    """
    # Stores the number of rows in the table
    first_row_of_table: int = 0
    last_row_of_table: int = 0

    # I added +1 to max_row because otherwise it will go upto the one to last but not the last row
    max_row: int = nz_xl.max_row + 1

    # Find the first row of the table after xl_row(see docs)
    for row_number in range(xl_row, max_row):
        value_of_cell_in_column_B = nz_xl.cell(
            row_number, 2
        )  # 2 corresponds to column B
        if value_of_cell_in_column_B.value is not None:
            first_row_of_table = row_number
            break

    # Find the last row of the table after xl_row(see docs)
    for row_number in range(first_row_of_table, max_row):
        value_of_cell_in_column_B = nz_xl.cell(
            row_number, 2
        )  # 2 corresponds to column B
        if value_of_cell_in_column_B.value is None:
            last_row_of_table = row_number
            break
        if row_number == max_row:
            last_row_of_table = max_row

    return last_row_of_table - first_row_of_table


def copy_coordinates_from_xl_to_rahshal(
    nz_xl, docx_table: Document, xl_row: int, table_count: int
) -> None:
    """
    The function `copy_coordinates_from_xl_to_rahshal` copies coordinates from an Excel file to a
    specified table in a Word document.

    Args:
    nz_xl: The variable `nz_xl` represents an Excel workbook or worksheet that contains the
    coordinates data.
    docx_table (Document): The `docx_table` parameter is a `Document` object representing a table in a
    Word document.
    xl_row (int): The `xl_row` parameter represents the row number in the Excel sheet from which the
    coordinates will be copied.
    table_count (int): The parameter `table_count` is used to determine which table in the document to
    copy the coordinates to. It is an integer value that represents the index of the table in the
    document. The tables in the document are numbered starting from 0.
    """
    # Start at row 5 because we need to leave space for the 5 default rows
    for row in range(5, len(docx_table.rows)):
        # explanation for the if statement is in xl_to_rahshal docstring
        if table_count == 0:
            for column in range(1, 7):  # Columns 1 to 7 in the excel
                cell = nz_xl.cell((xl_row + row) - 1, column + 1)
                # 'row + 2' because don't need to copy the first 2 rows
                if cell.value is not None:
                    docx_table.cell(row, column - 1).text = str(cell.value)
                    # 'column - 1' because in docx it start from index 0 and in excel it starts from index 1
        else:
            for column in range(1, 7):  # Columns 1 to 7 in the excel
                cell = nz_xl.cell((xl_row + row) - 1, column + 1)
                # 'row + 2' because don't need to copy the first 2 rows
                if cell.value is not None:
                    docx_table.cell(row, abs(column - 6)).text = str(cell.value)
                    # 'abs(column - 6)' explanation in the doc in the start of the document
